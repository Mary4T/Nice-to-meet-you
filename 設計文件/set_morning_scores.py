# -*- coding: utf-8 -*-
"""填入 morningScore 到夢境卡 Excel，然後重跑 import"""
import openpyxl

xlsx_path = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'

# morningScore 設計值：上/下/左/右
scores = {
    'dream_v1': [+2, -2, +1, +3],
    'dream_v2': [+1, +2, +3, +1],
    'dream_s1': [+2, -3, +2, +1],
    'dream_s2': [+1, -3, +3, +2],
    'dream_d1': [+2, +1, +3, +1],
    'dream_d2': [+2, -1, +2, +1],
    'dream_l1': [+2, -1, +3, +3],
    'dream_l2': [+2, -2, +2, +3],
    'dream_e1': [+2, +2, +3, -3],
    'dream_e2': [+3, -3, +2, +1],
}

wb = openpyxl.load_workbook(xlsx_path)
ws = wb['夢境卡']

headers = [cell.value for cell in ws[1]]
col_id    = headers.index('卡ID') + 1
col_score = headers.index('晨間分數') + 1

current_id = None
dir_idx = 0

for row in ws.iter_rows(min_row=2):
    card_id = row[col_id - 1].value
    if card_id:
        current_id = str(card_id).strip()
        dir_idx = 0
    if current_id in scores and dir_idx < 4:
        row[col_score - 1].value = scores[current_id][dir_idx]
        dir_idx += 1

wb.save(xlsx_path)
print('Done — morningScore filled')
