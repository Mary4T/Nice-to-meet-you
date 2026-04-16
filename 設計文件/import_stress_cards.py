# -*- coding: utf-8 -*-
"""
從 卡牌資料.xlsx 的「壓力牌」sheet 匯入壓力牌，
輸出至 src/phases/02_Lifetime/data/stress_cards.json
"""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX_PATH  = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
JSON_PATH  = r'D:\APP\Nice to meet you\.claude\worktrees\gracious-lamport\src\phases\02_Lifetime\data\stress_cards.json'

LEVEL_MAP = {'輕度': 'light', '中度': 'medium', '重度': 'heavy'}
POOL_MAP  = {'童年': 'childhood', '少年': 'juvenile', '青年': 'youth',
             '中年': 'midlife', '晚年': 'elder', '老年': 'elder', '': ''}
DIR_MAP   = {'上': 'up', '下': 'down', '左': 'left', '右': 'right'}
EFFECT_KEYS = ['聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性',
               '偽善度', '冷血度', '壓抑熵', '靈魂完整度']
EFFECT_EN = {
    '聲望': ('surface', 'Reputation'), '金錢': ('surface', 'Money'),
    '裏生命值': ('inner', 'Vitality'), '穩定': ('core', 'Stability'),
    '積極': ('core', 'Drive'), '理性': ('core', 'Logic'),
    '感性': ('core', 'Emotion'), '偽善度': ('soulToxins', 'Hypocrisy'),
    '冷血度': ('soulToxins', 'Ruthlessness'), '壓抑熵': ('soulToxins', 'Entropy'),
    '靈魂完整度': ('soul', 'SoulIntegrity'),
}

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['壓力牌']
headers = [cell.value for cell in ws[1]]
h = {name: idx for idx, name in enumerate(headers) if name}

cards = {}
current_id = None
current_card = None

for row in ws.iter_rows(min_row=2, values_only=True):
    card_id = str(row[h['卡ID']]).strip() if row[h['卡ID']] else ''
    direction_zh = str(row[h['方向']]).strip() if row[h['方向']] else ''
    if not direction_zh:
        continue
    direction = DIR_MAP.get(direction_zh)
    if not direction:
        continue

    # 新卡開始
    if card_id and card_id != current_id:
        current_id = card_id
        level_zh = str(row[h['壓力等級']]).strip() if row[h['壓力等級']] else ''
        pool_zh  = str(row[h['年齡池']]).strip() if row[h['年齡池']] else ''
        situation = str(row[h['情境']]).strip() if row[h['情境']] else ''
        current_card = {
            'id': current_id,
            'stressLevel': LEVEL_MAP.get(level_zh, 'light'),
            'agePool': POOL_MAP.get(pool_zh, pool_zh),
            'situation': situation,
        }
        cards[current_id] = current_card

    if current_card is None:
        continue

    # 效果
    effects: dict = {}
    for key in EFFECT_KEYS:
        if key not in h:
            continue
        val = row[h[key]]
        if val is None or val == '':
            continue
        try:
            val = int(val)
        except (ValueError, TypeError):
            continue
        if val == 0:
            continue
        group, en_key = EFFECT_EN[key]
        effects.setdefault(group, {})[en_key] = val

    option_text = str(row[h['選項文字']]).strip() if row[h['選項文字']] else ''
    current_card[direction] = {'text': option_text, 'effects': effects}

# 移除不完整的牌（缺少四個方向任一）
complete = {}
for cid, card in cards.items():
    if all(d in card for d in ('up', 'down', 'left', 'right')):
        complete[cid] = card
    else:
        missing = [d for d in ('up','down','left','right') if d not in card]
        print(f'WARNING: {cid} 缺少方向 {missing}，跳過')

output = {'stressCards': complete}
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'Done — 匯入 {len(complete)} 張壓力牌 → stress_cards.json')
