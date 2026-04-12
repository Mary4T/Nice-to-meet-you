# -*- coding: utf-8 -*-
"""在卡牌資料.xlsx 的「卡牌資料」sheet 末尾新增三張少年牌"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX_PATH = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.worksheets[0]

# 找最後一列
last_row = ws.max_row + 1

# 欄位順序（對照 import_cards.py）：
# 卡ID | 年齡池 | 情境 | 條件 | 方向 | 選項文字 | 聲望 | 金錢 | 裏生命值 | 穩定 | 積極 | 理性 | 感性 | 偽善度 | 冷血度 | 壓抑熵 | 靈魂完整度

CARDS = [
    # card_010 — 你喜歡的人
    {
        'id': 'card_010', 'pool': '少年',
        'situation': '你喜歡班上一個人。他不知道。你每天看著他，不知道要不要說。期末前你可能就不會再見到他了。',
        'dirs': [
            ('上', '鼓起勇氣，直接說出來',
             dict(穩定=-3, 積極=3, 感性=4)),
            ('下', '什麼都不說，默默放棄',
             dict(感性=-5, 壓抑熵=4)),
            ('左', '拜託共同朋友幫忙探一下口風',
             dict(理性=2, 感性=2, 偽善度=3)),
            ('右', '寫了一封信，但沒有署名',
             dict(積極=1, 感性=3, 偽善度=4)),
        ]
    },
    # card_011 — 你和別人不太一樣
    {
        'id': 'card_011', 'pool': '少年',
        'situation': '你不喜歡大家都在瘋的那個東西。電玩、球隊、那個流行歌手——你就是提不起勁。你開始覺得，是不是自己有什麼問題。',
        'dirs': [
            ('上', '假裝喜歡，努力跟上大家',
             dict(感性=-3, 偽善度=6)),
            ('下', '找到一個也格格不入的人，和他待在一起',
             dict(穩定=2, 感性=4)),
            ('左', '認真搞清楚自己到底喜歡什麼',
             dict(積極=5, 理性=3, 壓抑熵=-2)),
            ('右', '乾脆一個人，假裝不在乎',
             dict(穩定=-2, 積極=2, 冷血度=2, 壓抑熵=3)),
        ]
    },
    # card_012 — 家人發現你的秘密
    {
        'id': 'card_012', 'pool': '少年',
        'situation': '你有一件事沒讓家人知道。不是很大的事，但對你來說很重要。今天他們發現了。',
        'dirs': [
            ('上', '解釋清楚，讓他們真的了解',
             dict(穩定=4, 積極=2, 感性=3)),
            ('下', '輕描淡寫帶過，說沒什麼大不了',
             dict(穩定=-2, 偽善度=5)),
            ('左', '生氣說這是你自己的事，不需要他們管',
             dict(聲望=-2, 積極=3, 感性=-3)),
            ('右', '說謊，把事情圓回去',
             dict(穩定=-4, 偽善度=8, 壓抑熵=4)),
        ]
    },
]

COL_NAMES = ['卡ID', '年齡池', '情境', '條件', '方向', '選項文字',
             '聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性',
             '偽善度', '冷血度', '壓抑熵', '靈魂完整度']
EFFECT_COLS = {'聲望': 7, '金錢': 8, '裏生命值': 9, '穩定': 10, '積極': 11,
               '理性': 12, '感性': 13, '偽善度': 14, '冷血度': 15, '壓抑熵': 16, '靈魂完整度': 17}

row = last_row + 1  # 空一行
for card in CARDS:
    first = True
    for dir_zh, text, effects in card['dirs']:
        ws.cell(row, 1, card['id'] if first else None)
        ws.cell(row, 2, card['pool'] if first else None)
        ws.cell(row, 3, card['situation'] if first else None)
        ws.cell(row, 4, None)   # 條件
        ws.cell(row, 5, dir_zh)
        ws.cell(row, 6, text)
        for key, val in effects.items():
            ws.cell(row, EFFECT_COLS[key], val)
        first = False
        row += 1
    row += 1  # 空行分隔

wb.save(XLSX_PATH)
print('Done — card_010 / 011 / 012 已寫入 Excel')
