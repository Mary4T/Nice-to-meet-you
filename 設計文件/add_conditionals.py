# -*- coding: utf-8 -*-
"""
為 card_002（右）、card_004（下）、card_005（右）新增條件選項列
插入順序：由下往上，避免行號偏移
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX_PATH = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.worksheets[0]

# 欄位索引（1-based）
# 卡ID=1, 年齡池=2, 情境=3, 條件=4, 方向=5, 選項文字=6,
# 聲望=7, 金錢=8, 裏生命值=9, 穩定=10, 積極=11, 理性=12, 感性=13,
# 偽善度=14, 冷血度=15, 壓抑熵=16, 靈魂完整度=17

def insert_conditional(ws, after_row, direction, condition, text, vals):
    """在 after_row 之後插入一條件選項列"""
    ws.insert_rows(after_row + 1)
    r = after_row + 1
    ws.cell(r, 4).value = condition   # 條件
    ws.cell(r, 5).value = direction   # 方向
    ws.cell(r, 6).value = text        # 選項文字
    for col, val in vals.items():
        ws.cell(r, col).value = val

# ── 由下往上插入，避免行號偏移 ──────────────────────────────

# card_005 右（row 21）→ Ruthlessness>=50
insert_conditional(
    ws, after_row=21,
    direction='右',
    condition='Ruthlessness>=50',
    text='你微笑，在心裡開始計算他什麼時候會需要你。',
    vals={13: -10, 14: 8, 15: 10}   # 感性=-10, 偽善度=+8, 冷血度=+10
)
print('card_005 右 條件列已插入')

# card_004 下（row 15）→ Entropy>=70
insert_conditional(
    ws, after_row=15,
    direction='下',
    condition='Entropy>=70',
    text='你早就知道這個人不是你了。但已經沒差了。繼續',
    vals={10: -7, 14: 12, 16: 8}   # 穩定=-7, 偽善度=+12, 壓抑熵=+8
)
print('card_004 下 條件列已插入')

# card_002 右（row 9）→ Hypocrisy>=50
insert_conditional(
    ws, after_row=9,
    direction='右',
    condition='Hypocrisy>=50',
    text='答應。至於現在這份工作——讓他們自己想辦法',
    vals={8: 15, 10: -6, 14: 13, 15: 5}   # 金錢=+15, 穩定=-6, 偽善度=+13, 冷血度=+5
)
print('card_002 右 條件列已插入')

wb.save(XLSX_PATH)
print('Excel 已儲存')
