# -*- coding: utf-8 -*-
"""把 card_006~009 補進卡牌資料.xlsx 的「卡牌資料」sheet"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

xlsx_path = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.worksheets[0]  # 卡牌資料

headers = [cell.value for cell in ws[1]]
h = {name: i+1 for i, name in enumerate(headers) if name}

white_fill = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
gray_fill  = PatternFill('solid', start_color='F5F5F5', end_color='F5F5F5')
data_font  = Font(name='Arial')
wrap_align = Alignment(wrap_text=True, vertical='top')
top_align  = Alignment(vertical='top')

# 目前 Excel 最後一張牌的行數
existing_ids = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        existing_ids.append(str(row[0]))
# 用已有牌數決定交替底色起點
card_count_so_far = len(existing_ids)

missing_cards = [
    {
        'id': 'card_006', 'pool': '童年',
        'context': '你最好的朋友開始和另一群人混。他邀你一起，但那群人你不喜歡。',
        'options': [
            {'dir': '上', 'text': '硬著頭皮去，朋友比較重要',        '感性': 3,  '穩定': -2, '偽善度': 3},
            {'dir': '下', 'text': '直接說不喜歡那群人',               '積極': 3,  '感性': -2, '聲望': -1},
            {'dir': '左', 'text': '假裝很忙，慢慢疏遠',              '穩定': 1,  '偽善度': 6, '壓抑熵': 2},
            {'dir': '右', 'text': '試著和那群人相處看看，也許沒那麼差','理性': 3,  '感性': 1,  '聲望': 2},
        ]
    },
    {
        'id': 'card_007', 'pool': '童年',
        'context': '期末考沒考好。父母還不知道。成績單明天就要簽名交回去。',
        'options': [
            {'dir': '上', 'text': '主動告訴爸媽，先承認再說',          '穩定': 4,  '積極': 2},
            {'dir': '下', 'text': '自己簽名，能混就混',                '穩定': -3, '偽善度': 7, '壓抑熵': 4},
            {'dir': '左', 'text': '拜託哥哥姊姊幫忙說情',              '感性': 2,  '理性': 1,  '偽善度': 3},
            {'dir': '右', 'text': '把成績單藏起來，拖到他們忘記',      '穩定': -4, '偽善度': 5, '冷血度': 2, '壓抑熵': 5},
        ]
    },
    {
        'id': 'card_008', 'pool': '童年',
        'context': '同學的筆不見了。你看到另一個同學把它塞進書包——但那個人是班上最受歡迎的。',
        'options': [
            {'dir': '上', 'text': '當場說出來，不管後果',              '積極': 4,  '穩定': 2,  '聲望': -3},
            {'dir': '下', 'text': '假裝沒看見。不是你的事',            '感性': -4, '冷血度': 5, '壓抑熵': 3},
            {'dir': '左', 'text': '私下告訴失主，讓他自己決定',        '理性': 3,  '感性': 3},
            {'dir': '右', 'text': '悄悄告訴老師，不留名字',            '穩定': 3,  '偽善度': 2},
        ]
    },
    {
        'id': 'card_009', 'pool': '童年',
        'context': '你一直很想學畫畫。但父母說那沒有前途，要你去補數學。補習班已經報名了。',
        'options': [
            {'dir': '上', 'text': '去補數學。家人的話有道理',           '理性': 3,  '積極': -2, '感性': -2, '壓抑熵': 4},
            {'dir': '下', 'text': '鬧脾氣拒絕，堅持要學畫',            '積極': 5,  '穩定': -3, '聲望': -2},
            {'dir': '左', 'text': '表面答應，私下找時間偷練',           '積極': 3,  '偽善度': 5, '壓抑熵': 2},
            {'dir': '右', 'text': '認真和父母談，說明為什麼這對你重要', '積極': 4,  '理性': 3,  '感性': 2},
        ]
    },
]

stat_cols = ['聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性', '偽善度', '冷血度', '壓抑熵', '靈魂完整度']
current_row = ws.max_row + 1

for card_idx, card in enumerate(missing_cards):
    fill = white_fill if (card_count_so_far + card_idx) % 2 == 0 else gray_fill
    for opt_idx, opt in enumerate(card['options']):
        row_data = [None] * len(headers)
        if opt_idx == 0:
            row_data[h['卡ID'] - 1]   = card['id']
            row_data[h['年齡池'] - 1] = card['pool']
            row_data[h['情境'] - 1]   = card['context']
        row_data[h['方向'] - 1]      = opt['dir']
        row_data[h['選項文字'] - 1]  = opt['text']
        for stat in stat_cols:
            if stat in opt:
                row_data[h[stat] - 1] = opt[stat]
        ws.append(row_data)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = fill
            cell.font = data_font
            col_name = headers[col_idx - 1]
            if col_name in ('情境', '選項文字'):
                cell.alignment = wrap_align
            else:
                cell.alignment = top_align
        ws.row_dimensions[current_row].height = 45 if opt_idx == 0 else 30
        current_row += 1

wb.save(xlsx_path)
print('Done — card_006~009 added to Excel')
