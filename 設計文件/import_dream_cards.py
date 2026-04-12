# -*- coding: utf-8 -*-
"""
從 卡牌資料.xlsx 的「夢境卡」工作表，轉換成 dream_cards.json。
欄位對應：
  觸發維度  → triggerDimension
  夢境情境  → situation
  裏生命值  → core.Vitality
  穩定      → core.Stability
  積極      → core.Drive
  理性      → core.Logic
  感性      → core.Emotion
  偽善度    → soulToxins.Hypocrisy
  冷血度    → soulToxins.Ruthlessness
  壓抑熵    → soulToxins.Entropy
  晨間分數  → morningScore
"""
import json
import openpyxl

xlsx_path  = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
json_path  = r'D:\APP\Nice to meet you\src\phases\02_Lifetime\data\dream_cards.json'

dir_map = {'上': 'up', '下': 'down', '左': 'left', '右': 'right'}

core_map = {
    '裏生命值': 'Vitality',
    '穩定':     'Stability',
    '積極':     'Drive',
    '理性':     'Logic',
    '感性':     'Emotion',
}
toxin_map = {
    '偽善度': 'Hypocrisy',
    '冷血度': 'Ruthlessness',
    '壓抑熵': 'Entropy',
}

wb = openpyxl.load_workbook(xlsx_path)
if '夢境卡' not in wb.sheetnames:
    print('找不到「夢境卡」工作表')
    exit(1)

ws = wb['夢境卡']
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
h = {name: idx for idx, name in enumerate(headers) if name}

cards = {}
current_id = None
current_card = None

for row in rows[1:]:
    if all(v is None for v in row):
        continue

    card_id = row[h['卡ID']]
    if card_id:
        current_id = str(card_id).strip()
        dim        = str(row[h['觸發維度']]).strip() if row[h['觸發維度']] else ''
        situation  = str(row[h['夢境情境']]).strip() if row[h['夢境情境']] else ''
        current_card = {
            'id': current_id,
            'triggerDimension': dim,
            'situation': situation,
        }
        cards[current_id] = current_card

    if current_card is None:
        continue

    dir_zh = str(row[h['方向']]).strip() if row[h['方向']] else ''
    dir_en = dir_map.get(dir_zh)
    if not dir_en:
        continue

    text = str(row[h['選項文字']]).strip() if row[h['選項文字']] else ''
    core_effects = {}
    toxin_effects = {}

    for col, key in core_map.items():
        val = row[h[col]]
        if val is not None and val != '':
            core_effects[key] = int(val)

    for col, key in toxin_map.items():
        val = row[h[col]]
        if val is not None and val != '':
            toxin_effects[key] = int(val)

    morning_val = row[h['晨間分數']]
    morning_score = int(morning_val) if morning_val is not None and morning_val != '' else 0

    option = {'text': text}
    if core_effects:
        option['effects'] = {'core': core_effects}
    else:
        option['effects'] = {}
    if toxin_effects:
        option['effects']['soulToxins'] = toxin_effects
    option['morningScore'] = morning_score

    current_card[dir_en] = option

result = {'cards': cards}
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f'Done — {len(cards)} 張夢境卡寫入 dream_cards.json')
