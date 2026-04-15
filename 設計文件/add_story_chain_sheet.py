# -*- coding: utf-8 -*-
"""
在卡牌資料.xlsx 中：
1. 新增「故事鏈牌」sheet，填入測試用故事鏈資料
2. 在 card_031 的「上」選項填入後續事件 chain_001
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

XLSX_PATH = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
wb = openpyxl.load_workbook(XLSX_PATH)

# ── 1. 在 card_031「上」選項填後續事件 chain_001 ────────────────
ws_cards = wb['卡牌資料']
headers = [cell.value for cell in ws_cards[1]]
h = {name: idx + 1 for idx, name in enumerate(headers) if name}
COL_後續事件 = h['後續事件']

for row in ws_cards.iter_rows(min_row=2, values_only=False):
    card_id = row[h['卡ID'] - 1].value
    direction = row[h['方向'] - 1].value
    if card_id == 'card_031' and direction == '上':
        row[COL_後續事件 - 1].value = 'sc_001'
        print(f'card_031 上 → 後續事件 = sc_001')
        break

# ── 2. 新增「故事鏈牌」sheet ──────────────────────────────────
if '故事鏈牌' in wb.sheetnames:
    del wb['故事鏈牌']

ws = wb.create_sheet('故事鏈牌')

HEADERS = ['鏈ID', '卡ID', '情境', '方向', '選項文字', '下一張',
           '結束今日', '聲望', '金錢', '裏生命值', '穩定', '積極',
           '理性', '感性', '偽善度', '冷血度', '壓抑熵', '靈魂完整度']
EFFECT_COLS = {k: i + 8 for i, k in enumerate(
    ['聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性',
     '偽善度', '冷血度', '壓抑熵', '靈魂完整度'])}

# 標題列
header_fill = PatternFill('solid', fgColor='1F2937')
for col, name in enumerate(HEADERS, 1):
    cell = ws.cell(1, col, name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# 故事鏈牌資料
# chain_001：從 card_031「說實話」觸發
# sc_001 → sc_002 或 sc_003（依選擇）
# sc_002 / sc_003 → 鏈結束（下一張=空）

CHAIN_DATA = [
    # (鏈ID, 卡ID, 情境, 方向, 選項文字, 下一張, 結束今日, 效果dict)
    # 同一張牌的後續方向，鏈ID與卡ID留空（與普通牌慣例一致）
    ('chain_001', 'sc_001',
     '你把事情說清楚了。媽媽沒有立刻回應，沉默了一會兒。',
     '上', '繼續等她說話', 'sc_002', False, {}),
    ('', '', '', '下', '說你知道這樣不對', 'sc_002', False, {'穩定': 2}),
    ('', '', '', '左', '忍不住又道了一次歉', 'sc_002', False, {'壓抑熵': 2}),
    ('', '', '', '右', '問她是不是很失望', 'sc_003', False, {'感性': 2}),

    ('', 'sc_002',
     '媽媽說：「我知道你想要那個東西。但你應該來告訴我。」',
     '上', '說你以後會直接問', None, True, {'積極': 2, '穩定': 2}),
    ('', '', '', '下', '說對不起', None, True, {'穩定': 3, '壓抑熵': -2}),
    ('', '', '', '左', '點頭，什麼都沒說', None, True, {'壓抑熵': 2}),
    ('', '', '', '右', '問她能不能原諒你', None, True, {'感性': 3}),

    ('', 'sc_003',
     '她說：「我知道你有時候怕開口。但我希望你不要用這種方式。」',
     '上', '說你懂了', None, True, {'穩定': 3, '感性': 2}),
    ('', '', '', '下', '問她為什麼覺得你怕開口', None, True, {'感性': 4, '理性': 2}),
    ('', '', '', '左', '說你不知道為什麼沒說', None, True, {'壓抑熵': -3}),
    ('', '', '', '右', '沉默', None, True, {'壓抑熵': 3}),
]

for row_data in CHAIN_DATA:
    chain_id, card_id, situation, direction, text, next_card, ends_day, effects = row_data
    row = [chain_id, card_id, situation, direction, text,
           next_card if next_card else '',
           'TRUE' if ends_day else 'FALSE']
    # 效果欄（聲望到靈魂完整度）
    effect_order = ['聲望', '金錢', '裏生命值', '穩定', '積極', '理性',
                    '感性', '偽善度', '冷血度', '壓抑熵', '靈魂完整度']
    for k in effect_order:
        row.append(effects.get(k, None))
    ws.append(row)

# 欄寬
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 6
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10

wb.save(XLSX_PATH)
print(f'Done — 故事鏈牌 sheet 建立完成（{len(CHAIN_DATA)} 行）')
