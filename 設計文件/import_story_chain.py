# -*- coding: utf-8 -*-
"""
從 卡牌資料.xlsx 的「故事鏈牌」sheet 轉換成 story_chain_cards.json。

欄位對應：
  鏈ID      → chainId
  卡ID      → 牌的 key（如 sc_001）
  情境      → situation
  方向      → up/down/left/right
  選項文字  → text
  下一張    → next（空白 = null，鏈結束）
  結束今日  → endsDay（TRUE/FALSE）
  聲望～靈魂完整度 → effects（同普通牌）
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import json
import openpyxl

XLSX_PATH   = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
JSON_PATH   = r'D:\APP\Nice to meet you\src\phases\02_Lifetime\data\story_chain_cards.json'
JSON_PATH_WT = r'D:\APP\Nice to meet you\.claude\worktrees\gracious-lamport\src\phases\02_Lifetime\data\story_chain_cards.json'

DIR_MAP = {'上': 'up', '下': 'down', '左': 'left', '右': 'right'}

SURFACE_MAP = {'聲望': 'Reputation', '金錢': 'Money'}
CORE_MAP    = {'裏生命值': 'Vitality', '穩定': 'Stability', '積極': 'Drive',
               '理性': 'Logic', '感性': 'Emotion'}
TOXIN_MAP   = {'偽善度': 'Hypocrisy', '冷血度': 'Ruthlessness', '壓抑熵': 'Entropy'}

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['故事鏈牌']

headers = [cell.value for cell in ws[1]]
h = {name: idx + 1 for idx, name in enumerate(headers) if name}

def build_effects(row):
    effects = {}
    surface = {}
    for col, key in SURFACE_MAP.items():
        v = row[h[col] - 1] if col in h else None
        if v is not None and v != '': surface[key] = int(v)
    if surface: effects['surface'] = surface

    core = {}
    for col, key in CORE_MAP.items():
        v = row[h[col] - 1] if col in h else None
        if v is not None and v != '': core[key] = int(v)
    if core: effects['core'] = core

    toxins = {}
    for col, key in TOXIN_MAP.items():
        v = row[h[col] - 1] if col in h else None
        if v is not None and v != '': toxins[key] = int(v)
    if toxins: effects['soulToxins'] = toxins

    si = row[h['靈魂完整度'] - 1] if '靈魂完整度' in h else None
    if si is not None and si != '': effects['soulIntegrity'] = int(si)
    return effects

cards = {}
current_id = None
current_card = None

for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue

    card_id = str(row[h['卡ID'] - 1]).strip() if row[h['卡ID'] - 1] else None
    if card_id and card_id != current_id:
        current_id = card_id
        chain_id_val = row[h['鏈ID'] - 1] if row[h['鏈ID'] - 1] else None
        situation = str(row[h['情境'] - 1]).strip() if row[h['情境'] - 1] else ''
        current_card = {
            'id': current_id,
            'chainId': str(chain_id_val).strip() if chain_id_val else '',
            'situation': situation,
        }
        cards[current_id] = current_card

    if current_card is None:
        continue

    dir_zh = str(row[h['方向'] - 1]).strip() if row[h['方向'] - 1] else ''
    dir_en = DIR_MAP.get(dir_zh)
    if not dir_en:
        continue

    text     = str(row[h['選項文字'] - 1]).strip() if row[h['選項文字'] - 1] else ''
    next_raw = row[h['下一張'] - 1] if '下一張' in h else None
    next_id  = str(next_raw).strip() if next_raw and str(next_raw).strip() else None
    ends_raw = row[h['結束今日'] - 1] if '結束今日' in h else None
    ends_day = str(ends_raw).strip().upper() == 'TRUE' if ends_raw else False
    effects  = build_effects(row)

    current_card[dir_en] = {
        'text': text,
        'effects': effects,
        'next': next_id,
        'endsDay': ends_day,
    }

result = {'cards': cards}
for path in [JSON_PATH, JSON_PATH_WT]:
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

print(f'Done — {len(cards)} 張故事鏈牌寫入 story_chain_cards.json')
