# -*- coding: utf-8 -*-
"""
在卡牌資料.xlsx 新增「夢境卡」工作表，填入現有 dream_cards.json 資料。
欄位：卡ID / 觸發維度 / 夢境情境 / 方向 / 選項文字 /
      裏生命值 / 穩定 / 積極 / 理性 / 感性 / 偽善度 / 冷血度 / 壓抑熵 / 晨間分數
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

xlsx_path = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'

wb = load_workbook(xlsx_path)

# 如果已有「夢境卡」sheet 就刪掉重建
if '夢境卡' in wb.sheetnames:
    del wb['夢境卡']

ws = wb.create_sheet('夢境卡')

# ── 欄位定義 ──────────────────────────────────────────────
headers = [
    '卡ID', '觸發維度', '夢境情境', '方向', '選項文字',
    '裏生命值', '穩定', '積極', '理性', '感性',
    '偽善度', '冷血度', '壓抑熵', '晨間分數'
]

col_widths = {
    '卡ID': 12, '觸發維度': 14, '夢境情境': 42, '方向': 6, '選項文字': 32,
    '裏生命值': 10, '穩定': 10, '積極': 10, '理性': 10, '感性': 10,
    '偽善度': 10, '冷血度': 10, '壓抑熵': 10, '晨間分數': 10
}

header_fill = PatternFill('solid', start_color='FFD700', end_color='FFD700')
white_fill  = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
gray_fill   = PatternFill('solid', start_color='F0F4FF', end_color='F0F4FF')  # 淡藍，區分夢境卡
header_font = Font(name='Arial', bold=True, color='000000')
data_font   = Font(name='Arial')
wrap_align  = Alignment(wrap_text=True, vertical='top')
top_align   = Alignment(vertical='top')
center_top  = Alignment(horizontal='center', vertical='top')

ws.append(headers)
ws.row_dimensions[1].height = 20
ws.freeze_panes = 'A2'

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[h]

header_map = {h: i+1 for i, h in enumerate(headers)}

# ── 現有夢境卡資料（從 dream_cards.json 對應） ────────────
# core effects: Vitality=裏生命值, Stability=穩定, Drive=積極, Logic=理性, Emotion=感性
# soulToxins:   Entropy=壓抑熵, Hypocrisy=偽善度, Ruthlessness=冷血度
# morningScore: 晨間分數（暫填 0，待設計師調整）

dream_cards = [
    {
        'id': 'dream_v1', 'dim': 'Vitality',
        'situation': '夢裡你站在水邊。水面映出的不是你的臉，是另一個你——看起來更輕鬆、更完整。',
        'options': [
            {'dir': '上', 'text': '伸手觸碰水面',               '裏生命值': 5,  '壓抑熵': -8,  '晨間分數': 0},
            {'dir': '下', 'text': '轉身離開，不想看',            '裏生命值': -3, '壓抑熵': -4,  '晨間分數': 0},
            {'dir': '左', 'text': '在水邊坐下，靜靜看著那個影像','穩定': 3,      '壓抑熵': -6,  '晨間分數': 0},
            {'dir': '右', 'text': '對著水面說：「我知道了。」',  '裏生命值': 3,  '感性': 2, '壓抑熵': -9, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_v2', 'dim': 'Vitality',
        'situation': '夢裡你在奔跑，腿很重，但你繼續跑。你不確定是在追什麼，還是在逃什麼。',
        'options': [
            {'dir': '上', 'text': '跑得更快',              '積極': 4,  '壓抑熵': -5,  '晨間分數': 0},
            {'dir': '下', 'text': '停下來，喘口氣',         '穩定': 3,  '壓抑熵': -7,  '晨間分數': 0},
            {'dir': '左', 'text': '問自己：你在逃什麼？',    '理性': 3,  '壓抑熵': -8,  '晨間分數': 0},
            {'dir': '右', 'text': '往旁邊轉，換一條路',      '裏生命值': 4, '壓抑熵': -6, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_s1', 'dim': 'Stability',
        'situation': '夢裡你站在地板裂開的房間。裂縫很細，但一直在擴大。你不確定還能站多久。',
        'options': [
            {'dir': '上', 'text': '往裂縫靠近，仔細看它', '理性': 3,  '壓抑熵': -7,  '晨間分數': 0},
            {'dir': '下', 'text': '不動，等著看會怎樣',   '穩定': -2, '壓抑熵': -4,  '晨間分數': 0},
            {'dir': '左', 'text': '找最穩的地方站著',     '穩定': 4,  '壓抑熵': -6,  '晨間分數': 0},
            {'dir': '右', 'text': '大聲呼救',             '感性': 4,  '壓抑熵': -5,  '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_s2', 'dim': 'Stability',
        'situation': '夢裡你回到家，但每個房間都比你記得的更小。你越走越擠，幾乎喘不過氣。',
        'options': [
            {'dir': '上', 'text': '打開窗子，想讓空間透氣',         '積極': 3,  '壓抑熵': -6, '晨間分數': 0},
            {'dir': '下', 'text': '就這樣待在最小的角落',           '穩定': -3, '壓抑熵': -3, '晨間分數': 0},
            {'dir': '左', 'text': '一個一個找你小時候放的東西',     '感性': 5,  '壓抑熵': -8, '晨間分數': 0},
            {'dir': '右', 'text': '走出去，離開這棟房子',           '積極': 2,  '穩定': 2, '壓抑熵': -7, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_d1', 'dim': 'Drive',
        'situation': '夢裡有一條筆直的路，無限延伸。你站在起點，不知道盡頭有什麼。',
        'options': [
            {'dir': '上', 'text': '開始走，不管盡頭',           '積極': 5,  '壓抑熵': -7, '晨間分數': 0},
            {'dir': '下', 'text': '原地等等，也許路會告訴你',    '穩定': 2,  '壓抑熵': -5, '晨間分數': 0},
            {'dir': '左', 'text': '問自己：你想去哪裡？',        '理性': 4,  '壓抑熵': -9, '晨間分數': 0},
            {'dir': '右', 'text': '往路旁走，看看有沒有其他方向','感性': 3,  '壓抑熵': -6, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_d2', 'dim': 'Drive',
        'situation': '有人在夢裡遞給你一樣東西，你沒看清楚是什麼。他說：「這是你的。」',
        'options': [
            {'dir': '上', 'text': '接過來，先拿再說',       '積極': 3,  '壓抑熵': -6, '晨間分數': 0},
            {'dir': '下', 'text': '謝絕，說你不需要',       '穩定': 2,  '壓抑熵': -4, '晨間分數': 0},
            {'dir': '左', 'text': '問：「這是什麼？」',      '理性': 4,  '壓抑熵': -7, '晨間分數': 0},
            {'dir': '右', 'text': '看著它，但遲遲不動',      '感性': 3,  '壓抑熵': -5, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_l1', 'dim': 'Logic',
        'situation': '夢裡有一道題你一直解不開。它懸在空中，黑字白底，靜靜等你。',
        'options': [
            {'dir': '上', 'text': '繼續解，絕不放棄',           '理性': 5,  '積極': 2,  '壓抑熵': -7, '晨間分數': 0},
            {'dir': '下', 'text': '闔上眼睛不看它',             '穩定': 2,              '壓抑熵': -4, '晨間分數': 0},
            {'dir': '左', 'text': '試著改變解題方式',           '理性': 3,  '感性': 2,  '壓抑熵': -8, '晨間分數': 0},
            {'dir': '右', 'text': '說出聲：「我不知道答案。」', '穩定': 4,              '壓抑熵': -9, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_l2', 'dim': 'Logic',
        'situation': '夢裡你想說清楚一件事，但說出來的話和你想的不一樣。你試了很多次，都對不上。',
        'options': [
            {'dir': '上', 'text': '繼續試，一定有辦法說清楚',       '積極': 3,  '壓抑熵': -6, '晨間分數': 0},
            {'dir': '下', 'text': '不說了，讓對方自己去理解',       '感性': -3, '壓抑熵': -3, '晨間分數': 0},
            {'dir': '左', 'text': '換一種方式表達',                 '理性': 3,  '感性': 2,  '壓抑熵': -8, '晨間分數': 0},
            {'dir': '右', 'text': '問對方：「你聽到什麼了？」',     '感性': 5,  '壓抑熵': -9, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_e1', 'dim': 'Emotion',
        'situation': '夢裡有人在哭泣，聲音很近，但你找不到他在哪裡。',
        'options': [
            {'dir': '上', 'text': '繼續找，直到找到',              '積極': 3,  '感性': 3,  '壓抑熵': -8, '晨間分數': 0},
            {'dir': '下', 'text': '停下來，也跟著哭',              '感性': 5,              '壓抑熵': -7, '晨間分數': 0},
            {'dir': '左', 'text': '出聲說：「我在這裡。」',        '感性': 4,              '壓抑熵': -9, '晨間分數': 0},
            {'dir': '右', 'text': '告訴自己和自己沒關係，離開',    '冷血度': 3,            '壓抑熵': -3, '晨間分數': 0},
        ]
    },
    {
        'id': 'dream_e2', 'dim': 'Emotion',
        'situation': '夢裡你想對某個人說「對不起」，但對象的臉模糊了，你不確定是誰。',
        'options': [
            {'dir': '上', 'text': '說出來，不管對象是誰',                   '感性': 5,  '穩定': 2,  '壓抑熵': -10, '晨間分數': 0},
            {'dir': '下', 'text': '算了，也許說了也沒用',                   '偽善度': 4,            '壓抑熵': -4,  '晨間分數': 0},
            {'dir': '左', 'text': '試著想清楚，這個歉意是給誰的',           '理性': 3,  '感性': 3,  '壓抑熵': -8,  '晨間分數': 0},
            {'dir': '右', 'text': '只是站著，讓那份歉意留著',               '穩定': 3,              '壓抑熵': -7,  '晨間分數': 0},
        ]
    },
]

# ── 填入資料 ───────────────────────────────────────────────
stat_cols = ['裏生命值', '穩定', '積極', '理性', '感性', '偽善度', '冷血度', '壓抑熵', '晨間分數']
current_row = 2

for card_idx, card in enumerate(dream_cards):
    fill = white_fill if card_idx % 2 == 0 else gray_fill
    for opt_idx, opt in enumerate(card['options']):
        row_data = [None] * len(headers)
        if opt_idx == 0:
            row_data[header_map['卡ID'] - 1]      = card['id']
            row_data[header_map['觸發維度'] - 1]  = card['dim']
            row_data[header_map['夢境情境'] - 1]  = card['situation']
        row_data[header_map['方向'] - 1]      = opt['dir']
        row_data[header_map['選項文字'] - 1]  = opt['text']
        for stat in stat_cols:
            if stat in opt:
                row_data[header_map[stat] - 1] = opt[stat]
        ws.append(row_data)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = fill
            cell.font = data_font
            col_name = headers[col_idx - 1]
            if col_name in ('夢境情境', '選項文字'):
                cell.alignment = wrap_align
            elif col_name == '方向':
                cell.alignment = center_top
            else:
                cell.alignment = top_align
        ws.row_dimensions[current_row].height = 45 if opt_idx == 0 else 30
        current_row += 1

wb.save(xlsx_path)
print('Done — 夢境卡 sheet added to 卡牌資料.xlsx')
