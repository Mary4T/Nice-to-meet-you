# -*- coding: utf-8 -*-
"""
從 卡牌資料.xlsx 的「卡牌資料」sheet 轉換成 cards.json。
支援條件選項（conditionals）：
  條件欄格式：Vitality>30  /  Vitality>30&Entropy>50
  空白條件 = 基本選項
  同一方向多個條件列 → conditionals 陣列，Excel 由上到下順序排列
  優先順序：取最後一個所有條件都滿足的（代碼端由後往前找）

欄位對應：
  年齡池  → agePool  (童年/青年/中年/老年)
  情境    → situation
  條件    → condition string（空白=基本選項）
  方向    → up/down/left/right
  選項文字 → text
  聲望    → surface.Reputation
  金錢    → surface.Money
  裏生命值 → core.Vitality
  穩定    → core.Stability
  積極    → core.Drive
  理性    → core.Logic
  感性    → core.Emotion
  偽善度  → soulToxins.Hypocrisy
  冷血度  → soulToxins.Ruthlessness
  壓抑熵  → soulToxins.Entropy
  靈魂完整度 → soulIntegrity
"""
import json, re
import openpyxl

XLSX_PATH = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
JSON_PATH = r'D:\APP\Nice to meet you\src\phases\02_Lifetime\data\cards.json'
JSON_PATH_WT = r'D:\APP\Nice to meet you\.claude\worktrees\gracious-lamport\src\phases\02_Lifetime\data\cards.json'

POOL_MAP = {'童年': 'childhood', '少年': 'juvenile', '青年': 'youth', '中年': 'midlife', '老年': 'elder'}
DIR_MAP  = {'上': 'up', '下': 'down', '左': 'left', '右': 'right'}

SURFACE_MAP  = {'聲望': 'Reputation', '金錢': 'Money'}
CORE_MAP     = {'裏生命值': 'Vitality', '穩定': 'Stability', '積極': 'Drive', '理性': 'Logic', '感性': 'Emotion'}
TOXIN_MAP    = {'偽善度': 'Hypocrisy', '冷血度': 'Ruthlessness', '壓抑熵': 'Entropy'}

def parse_condition(cond_str):
    """'Vitality>30&Entropy>=50' → [{'key':'Vitality','op':'>','value':30}, ...]"""
    if not cond_str or str(cond_str).strip() == '':
        return None
    parts = str(cond_str).strip().split('&')
    result = []
    for p in parts:
        m = re.match(r'(\w+)(>=|<=|>|<|=)(-?\d+(?:\.\d+)?)', p.strip())
        if not m:
            raise ValueError(f'無法解析條件：{p}')
        key, op, val = m.group(1), m.group(2), m.group(3)
        result.append({'key': key, 'op': op, 'value': float(val) if '.' in val else int(val)})
    return result

def build_effects(row, h):
    effects = {}
    surface = {}
    for col, key in SURFACE_MAP.items():
        v = row[h[col] - 1] if col in h else None
        if v is not None and v != '': surface[key] = int(v)
    if surface: effects['surface'] = surface

    core = {}
    for col, key in CORE_MAP.items():
        v = row[h[col] - 1] if col in h else None
        if v is not None and v != '': core[key] = int(v)
    if core: effects['core'] = core

    toxins = {}
    for col, key in TOXIN_MAP.items():
        v = row[h[col] - 1] if col in h else None
        if v is not None and v != '': toxins[key] = int(v)
    if toxins: effects['soulToxins'] = toxins

    si_val = row[h['靈魂完整度'] - 1] if '靈魂完整度' in h else None
    if si_val is not None and si_val != '':
        effects['soulIntegrity'] = int(si_val)

    return effects

# ── 讀取 Excel ────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.worksheets[0]
headers = [cell.value for cell in ws[1]]
h = {name: idx+1 for idx, name in enumerate(headers) if name}

cards = {}
current_id = None
current_card = None

for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue

    card_id = row[h['卡ID'] - 1] if '卡ID' in h else None
    if card_id:
        current_id = str(card_id).strip()
        pool_zh = str(row[h['年齡池'] - 1]).strip() if row[h['年齡池'] - 1] else ''
        situation = str(row[h['情境'] - 1]).strip() if row[h['情境'] - 1] else ''
        current_card = {
            'id': current_id,
            'agePool': POOL_MAP.get(pool_zh, pool_zh),
            'situation': situation,
        }
        cards[current_id] = current_card

    if current_card is None:
        continue

    dir_zh  = str(row[h['方向'] - 1]).strip() if row[h['方向'] - 1] else ''
    dir_en  = DIR_MAP.get(dir_zh)
    if not dir_en:
        continue

    cond_raw  = row[h['條件'] - 1] if '條件' in h else None
    condition = parse_condition(cond_raw)
    text      = str(row[h['選項文字'] - 1]).strip() if row[h['選項文字'] - 1] else ''
    effects   = build_effects(row, h)

    if condition is None:
        # 基本選項
        current_card[dir_en] = {'text': text, 'effects': effects}
    else:
        # 條件選項：附加到該方向的 conditionals 陣列
        if dir_en not in current_card:
            current_card[dir_en] = {'text': '', 'effects': {}}
        if 'conditionals' not in current_card[dir_en]:
            current_card[dir_en]['conditionals'] = []
        current_card[dir_en]['conditionals'].append({
            'condition': condition,
            'text': text,
            'effects': effects,
        })

# ── 保留測試牌（不在 Excel 中，從各路徑讀回）──────────────
preserved_test_cards = {}
for try_path in [JSON_PATH_WT, JSON_PATH]:
    try:
        with open(try_path, encoding='utf-8') as f:
            old_data = json.load(f)
        for k, v in old_data.get('cards', {}).items():
            if k.startswith('__'):
                preserved_test_cards[k] = v
    except FileNotFoundError:
        continue
# 測試牌放在最前面
cards = {**preserved_test_cards, **cards}

result = {'cards': cards}
for path in [JSON_PATH, JSON_PATH_WT]:
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

print(f'Done — {len(cards)} 張牌（含測試牌）寫入 cards.json')
