# -*- coding: utf-8 -*-
"""
在卡牌資料.xlsx 新增「壓力牌」sheet，並填入 10 張壓力牌初稿。
欄位：壓力等級 | 年齡池 | 情境 | 方向 | 選項文字 | 聲望 | 金錢 | 裏生命值 | 穩定 | 積極 | 理性 | 感性 | 偽善度 | 冷血度 | 壓抑熵 | 靈魂完整度
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX_PATH = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
wb = openpyxl.load_workbook(XLSX_PATH)

# 若已有壓力牌 sheet 先刪除
if '壓力牌' in wb.sheetnames:
    del wb['壓力牌']

ws = wb.create_sheet('壓力牌')

HEADERS = ['卡ID', '壓力等級', '年齡池', '情境', '方向', '選項文字',
           '聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性',
           '偽善度', '冷血度', '壓抑熵', '靈魂完整度']

LEVEL_MAP = {'light': '輕度', 'medium': '中度', 'heavy': '重度'}
POOL_MAP  = {'childhood': '童年', 'youth': '青年', 'midlife': '中年', 'elder': '老年'}
DIR_MAP   = {'up': '上', 'down': '下', 'left': '左', 'right': '右'}

# ── 牌卡資料 ──────────────────────────────────────────────────
# 格式：(id, stressLevel, agePool, situation, {dir: (text, 聲望, 金錢, 裏生命值, 穩定, 積極, 理性, 感性, 偽善度, 冷血度, 壓抑熵, 靈魂完整度)})
CARDS = [
  ('stress_001', 'light', 'youth',
   '凌晨三點，你還醒著。明天有很重要的事。你的腦子停不下來，像一台沒有辦法關機的機器。', {
    'up':    ('硬逼自己睡，閉上眼睛數羊',                     0,0,0,-1,0,0,0,0,0,0,0),
    'down':  ('滑手機，等睏意自己來',                         0,0,0,0,-2,0,0,0,0,3,0),
    'left':  ('起來走走，把燈全關掉，坐在黑暗裡',             0,0,0,-2,0,0,-1,0,0,0,0),
    'right': ('把明天要做的事全列出來，試著清空腦袋',         0,0,0,0,-1,1,0,0,0,0,0),
  }),
  ('stress_002', 'light', 'midlife',
   '朋友傳來一張照片。他看起來過得很好。你盯著手機，說不清楚自己在想什麼，只是一直滑，一直看。', {
    'up':    ('按讚，然後把手機放下',                         0,0,0,0,0,0,0,4,0,0,0),
    'down':  ('傳訊熱情恭喜他，語氣比你想的還誇張',           0,0,0,0,0,0,-3,6,0,0,0),
    'left':  ('關掉頁面，不去想',                             0,0,0,-2,0,0,0,0,0,3,0),
    'right': ('認真問自己，你現在真正想要的是什麼',           0,0,0,0,-2,2,0,0,0,-2,0),
  }),
  ('stress_003', 'light', 'youth',
   '你說了一句話。說完就知道不對了。對方的臉色變了一下，然後裝作沒事。你們繼續說話，但什麼東西已經不一樣了。', {
    'up':    ('立刻道歉，直接說你說錯了',                     -1,0,0,2,0,0,0,0,0,0,0),
    'down':  ('裝作沒發現，繼續說別的',                       0,0,0,0,0,0,0,5,0,3,0),
    'left':  ('沉默，希望這件事自己消失',                     0,0,0,-3,0,0,0,0,0,4,0),
    'right': ('把話說清楚，就算尷尬也好',                     0,0,0,0,2,0,-2,0,0,0,0),
  }),
  ('stress_004', 'medium', 'youth',
   '你對便利商店店員凶了一句。他只是動作慢了一點。你走出去，站在外面，說不清楚剛才是怎麼了。你一直以為自己不是這種人。', {
    'up':    ('回去道歉',                                     -1,0,0,2,0,0,3,0,0,0,0),
    'down':  ('繼續走，告訴自己他本來就慢',                   0,0,0,0,0,0,0,7,5,0,0),
    'left':  ('站在外面，把這件事壓下去',                     0,0,0,-2,0,0,0,0,0,5,0),
    'right': ('打給一個人，說你現在狀態不好',                 0,0,0,-1,0,0,3,0,0,0,0),
  }),
  ('stress_005', 'medium', 'midlife',
   '鑰匙找不到。就是這樣。你找了三分鐘，然後坐在地上，開始哭。你知道這不是鑰匙的問題。你也不知道是什麼的問題。', {
    'up':    ('哭完，深呼吸，站起來繼續找',                   0,0,-3,0,0,0,0,0,0,-3,0),
    'down':  ('打電話請人幫忙，說自己不舒服',                 -2,0,0,0,0,0,2,0,0,0,0),
    'left':  ('坐著，讓自己慢慢平靜',                         0,0,0,-4,0,0,0,0,0,-3,0),
    'right': ('把手邊的東西推倒，讓情緒出來',                 0,0,0,-5,0,0,0,0,3,-2,0),
  }),
  ('stress_006', 'medium', 'youth',
   '主管問你一個問題。你知道你知道答案。但腦子裡什麼都沒有。空白，徹底的空白。所有人都在等。', {
    'up':    ('說「讓我想一下」，撐過這五秒',                 0,0,0,-2,0,-1,0,0,0,0,0),
    'down':  ('胡亂說了什麼，希望能帶過去',                   0,0,0,-3,0,0,0,6,0,0,0),
    'left':  ('老實說你現在腦袋不太轉',                       -3,0,0,0,0,0,2,0,0,0,0),
    'right': ('請對方再說一次問題，爭取時間',                 -1,0,0,0,0,1,0,0,0,0,0),
  }),
  ('stress_007', 'medium', 'midlife',
   '你想不起來上次真心笑是什麼時候了。不是應酬的笑，不是苦笑。是真的，從裡面出來的那種好。', {
    'up':    ('試著想一件以前喜歡做的事，這週去做',           0,0,0,0,3,0,0,0,0,-3,0),
    'down':  ('告訴自己這就是長大的代價，繼續',               0,0,0,0,0,0,0,5,0,6,0),
    'left':  ('跟一個信任的人說這件事',                       0,0,0,-1,0,0,3,0,0,0,0),
    'right': ('什麼都不做，繼續等',                           0,0,-4,0,0,0,0,0,0,5,0),
  }),
  ('stress_008', 'heavy', 'midlife',
   '你說了「好」。你不知道為什麼，嘴就開了。你看著自己點頭，像在看別人。等你回過神，已經太晚了。你甚至不確定你想拒絕的是這件事，還是整個現在的生活。', {
    'up':    ('馬上說你說錯了，你做不到',                     -4,0,0,3,0,0,0,0,0,0,0),
    'down':  ('接受，想辦法撐過去',                           0,0,0,0,-5,0,0,0,3,7,0),
    'left':  ('之後再找理由推掉，先應付過去',                 0,0,0,-3,0,0,0,9,0,0,0),
    'right': ('什麼都不說，讓這件事爛在心裡',                 0,0,-5,0,0,0,0,0,0,8,0),
  }),
  ('stress_009', 'heavy', 'midlife',
   '不是想死。只是想消失一陣子。不想被找到，不想回應任何人，不想是自己。就這樣。你不知道這個念頭從哪裡來的，也不知道什麼時候開始的。', {
    'up':    ('請幾天假，讓自己真的消失一下',                 0,-8,0,2,0,0,0,0,0,-5,0),
    'down':  ('繼續撐，讓這個念頭沉下去',                     0,0,-6,0,0,0,0,0,0,9,0),
    'left':  ('告訴一個人你現在的狀態',                       0,0,0,-2,0,0,4,0,0,-3,0),
    'right': ('去一個沒有人認識你的地方，坐著',               0,-5,0,-3,0,0,0,0,0,-4,0),
  }),
  ('stress_010', 'heavy', 'elder',
   '有人問你想要什麼。你張口，然後閉上。不是不知道怎麼說，是真的不知道。你已經很久沒問過自己這個問題了。你甚至不確定你還有資格想要什麼。', {
    'up':    ('說「我需要時間想」，然後真的去想',              0,0,0,-4,2,0,0,0,0,-3,0),
    'down':  ('說了一個聽起來合理的答案，但那不是你的',       0,0,-4,0,0,0,0,10,0,0,0),
    'left':  ('反問對方，用問題擋問題',                       0,0,0,0,0,1,0,5,0,0,0),
    'right': ('沉默很久，然後說「我不知道」',                 0,0,0,-5,0,0,5,0,0,-4,0),
  }),
]

# ── 寫入 sheet ────────────────────────────────────────────────
# Header 樣式
header_fill = PatternFill('solid', fgColor='2C3E6B')
header_font = Font(bold=True, color='FFFFFF', name='新細明體')
border_side = Side(style='thin', color='CCCCCC')
border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(1, col, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')

row = 2
for (card_id, level, pool, situation, dirs) in CARDS:
    first = True
    for dir_en in ['up', 'down', 'left', 'right']:
        vals = dirs[dir_en]
        text, rep, mon, vit, sta, drv, log, emo, hyp, rth, ent, si = (vals[0],) + vals[1:]
        ws.cell(row, 1, card_id if first else None)   # 卡ID
        ws.cell(row, 2, LEVEL_MAP[level] if first else None)  # 壓力等級
        ws.cell(row, 3, POOL_MAP[pool] if first else None)    # 年齡池
        ws.cell(row, 4, situation if first else None)          # 情境
        ws.cell(row, 5, DIR_MAP[dir_en])                       # 方向
        ws.cell(row, 6, text)                                  # 選項文字
        for ci, v in enumerate([rep, mon, vit, sta, drv, log, emo, hyp, rth, ent, si], 7):
            ws.cell(row, ci, v if v != 0 else None)
        for col in range(1, len(HEADERS)+1):
            ws.cell(row, col).border = border
        first = False
        row += 1
    # 空行分隔
    row += 1

# 欄寬
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 40

wb.save(XLSX_PATH)
print(f'Done — 壓力牌 sheet 已建立，共 {len(CARDS)} 張牌')
