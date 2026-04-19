# -*- coding: utf-8 -*-
"""
建立五個年齡池 Excel 並從卡牌資料.xlsx 遷移現有資料
每個 Excel 有三個分頁：一般牌卡、夢境卡、故事鏈
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC_PATH = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
OUT_DIR  = r'D:\APP\Nice to meet you\設計文件\劇本-正式'

POOLS = [
    ('童年', 'childhood', '童年'),
    ('少年', 'juvenile',  '少年'),
    ('青年', 'youth',     '青年'),
    ('中年', 'midlife',   '中年'),
    ('晚年', 'elder',     ['老年', '晚年']),  # 舊資料可能叫老年
]

# 舊方向→新方向對照（遷移用）
OLD_TO_NEW_DIR = {'上': 'A', '下': 'B', '左': 'C', '右': 'D'}

# ── 樣式 ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
CENTER      = Alignment(horizontal='center', vertical='center')

def style_header(ws, headers):
    for col, name in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 20

# ── 一般牌卡欄位定義 ──────────────────────────────────────────
# 【標籤】移到最前面；移除【靈魂標記】；新增【設計邏輯】在最後
# 方向欄：A/B/C/D（取代舊的上/下/左/右）
CARD_HEADERS = [
    '標籤',          # 移到最前，供劇本家分類管理
    '卡ID', '情境',
    '出現條件',      # 整張牌的出現條件，格式：Stability>30&day>=20
    '觸發機率',      # 0~100，空白=100%
    '方向',          # A / B / C / D（UI 隨機分配到上下左右）
    '選項文字', '條件',
    '聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性',
    '偽善度', '冷血度', '壓抑熵', '靈魂完整度',
    '後續事件',
    '夢數值',        # 選此選項時累加的夢數值
    '連結夢境卡ID',  # 選此選項後加入夢境卡池的夢境卡ID
    '設計邏輯',      # 純筆記欄，不匯入程式
]

# ── 夢境卡欄位定義 ────────────────────────────────────────────
DREAM_HEADERS = [
    '卡ID', '夢境情境', '方向', '選項文字',
    '裏生命值', '穩定', '積極', '理性', '感性',
    '偽善度', '冷血度', '壓抑熵', '靈魂完整度',
]

# ── 故事鏈欄位定義 ────────────────────────────────────────────
CHAIN_HEADERS = [
    '鏈ID', '卡ID', '情境', '方向', '選項文字', '下一張', '結束今日',
    '聲望', '金錢', '裏生命值', '穩定', '積極', '理性', '感性',
    '偽善度', '冷血度', '壓抑熵', '靈魂完整度',
]

# ── 讀取來源資料 ──────────────────────────────────────────────
src = openpyxl.load_workbook(SRC_PATH)

def get_rows(ws):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return headers, rows

src_headers_card, src_rows_card = get_rows(src['卡牌資料'])
src_headers_dream, src_rows_dream = get_rows(src['夢境卡'])
src_headers_chain, src_rows_chain = get_rows(src['故事鏈牌'])

def col_width(ws, widths):
    cols = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for i, w in enumerate(widths):
        ws.column_dimensions[cols[i]].width = w

# ── 為每個年齡池建立 Excel ─────────────────────────────────────
for pool_name, pool_en, pool_match in POOLS:
    match_values = pool_match if isinstance(pool_match, list) else [pool_match]
    wb = openpyxl.Workbook()

    # ── 分頁1：一般牌卡 ───────────────────────────────────────
    ws_card = wb.active
    ws_card.title = '一般牌卡'
    style_header(ws_card, CARD_HEADERS)
    # 23 欄的寬度（與 CARD_HEADERS 對齊）
    col_width(ws_card, [15, 12, 40, 20, 8, 5, 24, 20,
                         6, 6, 8, 6, 6, 6, 6, 6, 6, 6, 8,
                         12, 6, 14, 30])

    current_pool = None
    in_target_pool = False
    for row in src_rows_card:
        if row.get('年齡池'):
            current_pool = row['年齡池']
            in_target_pool = current_pool in match_values
        if not in_target_pool:
            continue
        # 方向值：舊 上/下/左/右 → 新 A/B/C/D
        old_dir = str(row.get('方向') or '')
        new_dir = OLD_TO_NEW_DIR.get(old_dir, old_dir)
        new_row = [
            None,       # 標籤（新欄，留空）
            row.get('卡ID'), row.get('情境'),
            None,       # 出現條件（新欄，留空）
            None,       # 觸發機率（新欄，留空）
            new_dir,    # 方向：A/B/C/D
            row.get('選項文字'), row.get('條件'),
            row.get('聲望'), row.get('金錢'), row.get('裏生命值'),
            row.get('穩定'), row.get('積極'), row.get('理性'), row.get('感性'),
            row.get('偽善度'), row.get('冷血度'), row.get('壓抑熵'), row.get('靈魂完整度'),
            row.get('後續事件'),
            None,       # 夢數值（新欄，留空）
            None,       # 連結夢境卡ID（新欄，留空）
            None,       # 設計邏輯（新欄，留空）
        ]
        ws_card.append(new_row)

    # ── 分頁2：夢境卡 ─────────────────────────────────────────
    ws_dream = wb.create_sheet('夢境卡')
    style_header(ws_dream, DREAM_HEADERS)
    col_width(ws_dream, [12, 40, 5, 24, 8, 6, 6, 6, 6, 6, 6, 6, 8])

    DIMENSION_POOL = {
        'Stability': '童年', 'Drive': '少年', 'Logic': '青年',
        'Emotion': '中年', 'Vitality': '晚年',
    }
    for row in src_rows_dream:
        dim = row.get('觸發維度', '')
        if DIMENSION_POOL.get(dim, '童年') == pool_name or (dim not in DIMENSION_POOL and pool_name == '童年'):
            new_row = [
                row.get('卡ID'), row.get('夢境情境'),
                row.get('方向'), row.get('選項文字'),
                row.get('裏生命值'), row.get('穩定'), row.get('積極'),
                row.get('理性'), row.get('感性'),
                row.get('偽善度'), row.get('冷血度'), row.get('壓抑熵'),
                None,   # 靈魂完整度（新欄）
            ]
            ws_dream.append(new_row)

    # ── 分頁3：故事鏈 ─────────────────────────────────────────
    ws_chain = wb.create_sheet('故事鏈')
    style_header(ws_chain, CHAIN_HEADERS)
    col_width(ws_chain, [12, 10, 40, 6, 24, 10, 8,
                          6, 6, 8, 6, 6, 6, 6, 6, 6, 6, 8])

    if pool_name == '童年':
        for row in src_rows_chain:
            new_row = [
                row.get('鏈ID'), row.get('卡ID'), row.get('情境'),
                row.get('方向'), row.get('選項文字'), row.get('下一張'), row.get('結束今日'),
                row.get('聲望'), row.get('金錢'), row.get('裏生命值'),
                row.get('穩定'), row.get('積極'), row.get('理性'), row.get('感性'),
                row.get('偽善度'), row.get('冷血度'), row.get('壓抑熵'), row.get('靈魂完整度'),
            ]
            ws_chain.append(new_row)

    # ── 填寫說明分頁 ───────────────────────────────────────────
    ws_note = wb.create_sheet('填寫說明')
    notes = [
        ['欄位', '說明'],
        ['標籤', '家庭/感情/健康/工作等，逗號分隔，供劇本家管理用'],
        ['出現條件', '整張牌的出現條件，格式同選項條件欄（如 Stability>30、day>=20）。空白=無條件限制'],
        ['觸發機率', '0~100 的整數，代表每天抽牌時此牌被抽到的機率。空白=100%'],
        ['方向', 'A / B / C / D（對應四個選項）。UI 每次隨機分配到上下左右方位，選項標籤固定為 ABCD'],
        ['夢數值', '選此選項時累加的夢數值（整數）。累積到10當天觸發夢境。空白=0'],
        ['連結夢境卡ID', '選此選項後加入夢境卡池的夢境卡ID（如 dream_s1）。空白=不連結'],
        ['設計邏輯', '純筆記欄，不匯入程式，供劇本家記錄設計意圖'],
        ['', ''],
        ['出現條件格式範例', ''],
        ['Stability>30', '穩定>30 時出現'],
        ['day>=20', '第20天以後出現'],
        ['Entropy>=80', '壓抑熵>=80 時出現'],
        ['Stability>30&Entropy<50', 'AND 條件，多條件用 & 連接'],
    ]
    for r in notes:
        ws_note.append(r)
    ws_note.column_dimensions['A'].width = 20
    ws_note.column_dimensions['B'].width = 55

    out_path = f'{OUT_DIR}\\{pool_name}.xlsx'
    wb.save(out_path)
    print(f'建立完成：{pool_name}.xlsx')

print('\nDone — 五個年齡池 Excel 建立完成')
