# -*- coding: utf-8 -*-
"""
將童年、少年、青年、中年 Excel 同步為晚年的新欄位結構
同時修正晚年夢境卡「方向」→「選項」

執行方式：python migrate_to_new_structure.py
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

DIR = r'D:\APP\Nice to meet you\設計文件\劇本-正式'

# 上下左右 → A/B/C/D（舊資料遷移）
DIR_TO_ABCD = {'上': 'A', '下': 'B', '左': 'C', '右': 'D'}

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
CENTER      = Alignment(horizontal='center', vertical='center')

def style_header(ws, headers, col_widths):
    """套用標題列樣式並設定欄寬"""
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, name)
        cell.font  = copy(HEADER_FONT)
        cell.fill  = copy(HEADER_FILL)
        cell.alignment = copy(CENTER)
    ws.row_dimensions[1].height = 20
    col_letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for i, w in enumerate(col_widths):
        ws.column_dimensions[col_letters[i]].width = w

def get_row_dict(ws):
    """取得所有資料列（dict 格式）"""
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows

# ══════════════════════════════════════════════════════════════
# 目標欄位定義（對應晚年結構）
# ══════════════════════════════════════════════════════════════

# 一般牌卡：(新欄位名, 舊欄位名 or None)
# None = 新增空白欄；舊欄位名 = 從舊欄複製資料
CARD_COLS = [
    ('卡ID',         '卡ID'),
    ('標籤',         '標籤'),
    ('情境',         '情境'),
    ('觸發條件1',    '出現條件'),   # 舊「出現條件」遷移至觸發條件1
    ('觸發條件2',    None),
    ('觸發機率',     '觸發機率'),
    ('選項',         '方向'),       # 舊「方向」改名；值同步轉換 上下左右→ABCD
    ('選項文字',     '選項文字'),
    ('選項出現條件', '條件'),       # 舊「條件」改名
    ('聲望',         '聲望'),
    ('金錢',         '金錢'),
    ('裏生命值',     '裏生命值'),
    ('穩定',         '穩定'),
    ('積極',         '積極'),
    ('理性',         '理性'),
    ('感性',         '感性'),
    ('偽善度',       '偽善度'),
    ('冷血度',       '冷血度'),
    ('壓抑熵',       '壓抑熵'),
    ('靈魂完整度',   '靈魂完整度'),
    ('後續事件',     '後續事件'),
    ('夢數值',       '夢數值'),
    ('連結夢境卡ID', '連結夢境卡ID'),
    ('設計邏輯',     None),         # 新增空白欄
]
CARD_WIDTHS = [12, 15, 40, 20, 20, 8, 6, 24, 22,
               6, 6, 8, 6, 6, 6, 6, 6, 6, 6, 8,
               12, 6, 14, 30]

# 夢境卡：(新欄位名, 舊欄位名 or None)
DREAM_COLS = [
    ('卡ID',       '卡ID'),
    ('夢境情境',   '夢境情境'),
    ('選項',       '方向'),        # 方向 → 選項
    ('選項文字',   '選項文字'),
    ('裏生命值',   '裏生命值'),
    ('穩定',       '穩定'),
    ('積極',       '積極'),
    ('理性',       '理性'),
    ('感性',       '感性'),
    ('偽善度',     '偽善度'),
    ('冷血度',     '冷血度'),
    ('壓抑熵',     '壓抑熵'),
    ('靈魂完整度', '靈魂完整度'),
    ('後續事件',   '後續事件'),   # 若舊檔無此欄則留空
]
DREAM_WIDTHS = [12, 40, 6, 24, 8, 6, 6, 6, 6, 6, 6, 6, 8, 12]

# 故事鏈：(新欄位名, 舊欄位名 or None)
CHAIN_COLS = [
    ('鏈ID',         '鏈ID'),
    ('卡ID',         '卡ID'),
    ('標籤',         None),         # 新增
    ('情境',         '情境'),
    ('選項',         '方向'),       # 方向 → 選項
    ('觸發條件1',    None),         # 新增
    ('觸發條件2',    None),         # 新增
    ('選項文字',     '選項文字'),
    ('選項出現條件', '條件'),       # 舊「條件」（若無則空白）
    ('下一張',       '下一張'),
    ('結束今日',     '結束今日'),
    ('聲望',         '聲望'),
    ('金錢',         '金錢'),
    ('裏生命值',     '裏生命值'),
    ('穩定',         '穩定'),
    ('積極',         '積極'),
    ('理性',         '理性'),
    ('感性',         '感性'),
    ('偽善度',       '偽善度'),
    ('冷血度',       '冷血度'),
    ('壓抑熵',       '壓抑熵'),
    ('靈魂完整度',   '靈魂完整度'),
    ('後續事件',     None),         # 新增
    ('夢數值',       None),         # 新增
    ('連結夢境卡ID', None),         # 新增
    ('設計邏輯',     None),         # 新增
]
CHAIN_WIDTHS = [12, 10, 15, 40, 6, 20, 20, 24, 22,
                10, 8, 6, 6, 8, 6, 6, 6, 6, 6, 6, 6, 8,
                12, 6, 14, 30]

# 填寫說明
NOTE_ROWS = [
    ['欄位',         '說明'],
    ['卡ID',         '每張牌唯一識別碼，格式：card_001'],
    ['標籤',         '家庭/感情/健康/工作等，逗號分隔，供劇本家管理用'],
    ['觸發條件1',    '整張牌的出現條件（第一條）。格式：Stability>30、day>=20、Entropy>=80。空白=無條件限制'],
    ['觸發條件2',    '整張牌的出現條件（第二條）。兩條皆填時需「同時成立（AND）」才能出牌。空白=忽略此條'],
    ['觸發機率',     '0~100 的整數，代表每天抽牌時此牌被抽到的機率。空白=100%'],
    ['選項',         'A / B / C / D（對應四個選項）。UI 每次隨機分配到上下左右方位'],
    ['選項文字',     '玩家看到的選項內容'],
    ['選項出現條件', '此選項的覆蓋條件。格式同觸發條件。成立時此選項取代基本選項。空白=永遠顯示'],
    ['下一張',       '（故事鏈用）下一張鏈內牌 ID；空白=鏈結束，回到主牌池'],
    ['結束今日',     '（故事鏈用）此選項是否結束今日。填 TRUE 或空白（FALSE）'],
    ['聲望',         '正數=增加，負數=減少'],
    ['金錢',         '正數=增加，負數=減少'],
    ['裏生命值',     '正數=恢復，負數=損耗'],
    ['穩定',         '裏層數值，正負皆可'],
    ['積極',         '裏層數值，正負皆可'],
    ['理性',         '裏層數值，正負皆可'],
    ['感性',         '裏層數值，正負皆可'],
    ['偽善度',       '靈魂毒素，正負皆可'],
    ['冷血度',       '靈魂毒素，正負皆可'],
    ['壓抑熵',       '靈魂毒素，正負皆可'],
    ['靈魂完整度',   '靈魂完整度，正負皆可'],
    ['後續事件',     '填故事鏈卡ID（如 chain_001_1），觸發故事鏈；或留空'],
    ['夢數值',       '選此選項時累加的夢數值（整數）。累積到10觸發夢境。空白=0'],
    ['連結夢境卡ID', '選此選項後加入夢境卡池的夢境卡ID（如 dream_s1）。空白=不連結'],
    ['設計邏輯',     '純筆記欄，不匯入程式，供劇本家記錄設計意圖'],
    ['',             ''],
    ['觸發條件格式範例', ''],
    ['Stability>30',      '穩定 > 30 時出牌'],
    ['day>=20',           '第 20 天以後出牌'],
    ['Entropy>=80',       '壓抑熵 >= 80 時出牌'],
    ['觸發條件1 + 觸發條件2', '兩欄皆填 → 兩個條件須同時成立才出牌（AND）'],
]

def migrate_sheet(old_ws, new_ws, col_map, col_widths):
    """依照 col_map 重建工作表欄位與資料"""
    new_headers = [new_col for new_col, _ in col_map]
    old_rows = get_row_dict(old_ws)

    style_header(new_ws, new_headers, col_widths)

    for old_row in old_rows:
        new_row = []
        for new_col, old_col in col_map:
            val = old_row.get(old_col) if old_col else None
            # 方向欄：上下左右 → ABCD
            if new_col == '選項' and val:
                val = DIR_TO_ABCD.get(str(val), val)
            new_row.append(val)
        new_ws.append(new_row)

def migrate_notes(ws):
    """重建填寫說明"""
    ws.delete_rows(1, ws.max_row)
    for row in NOTE_ROWS:
        ws.append(row)
    # 標題列樣式
    for col in [1, 2]:
        cell = ws.cell(1, col)
        cell.font  = copy(HEADER_FONT)
        cell.fill  = copy(HEADER_FILL)
        cell.alignment = copy(CENTER)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 58

# ══════════════════════════════════════════════════════════════
# 執行遷移：童年、少年、青年、中年
# ══════════════════════════════════════════════════════════════
for pool in ['童年', '少年', '青年', '中年']:
    path = DIR + '\\' + pool + '.xlsx'
    old_wb = openpyxl.load_workbook(path)
    new_wb = openpyxl.Workbook()

    # 一般牌卡
    new_card_ws = new_wb.active
    new_card_ws.title = '一般牌卡'
    migrate_sheet(old_wb['一般牌卡'], new_card_ws, CARD_COLS, CARD_WIDTHS)

    # 夢境卡
    new_dream_ws = new_wb.create_sheet('夢境卡')
    migrate_sheet(old_wb['夢境卡'], new_dream_ws, DREAM_COLS, DREAM_WIDTHS)

    # 故事鏈
    new_chain_ws = new_wb.create_sheet('故事鏈')
    migrate_sheet(old_wb['故事鏈'], new_chain_ws, CHAIN_COLS, CHAIN_WIDTHS)

    # 填寫說明
    new_note_ws = new_wb.create_sheet('填寫說明')
    migrate_notes(new_note_ws)

    new_wb.save(path)
    print(pool + '.xlsx 遷移完成')

# ══════════════════════════════════════════════════════════════
# 修正晚年：夢境卡「方向」→「選項」；更新填寫說明
# ══════════════════════════════════════════════════════════════
path_elder = DIR + '\\晚年.xlsx'
wb_elder = openpyxl.load_workbook(path_elder)

# 夢境卡：將標題列「方向」改為「選項」
ws_dream = wb_elder['夢境卡']
for cell in ws_dream[1]:
    if cell.value == '方向':
        cell.value = '選項'
        break

# 填寫說明：整體重建
ws_note = wb_elder['填寫說明']
ws_note.delete_rows(1, ws_note.max_row)
for row in NOTE_ROWS:
    ws_note.append(row)
for col in [1, 2]:
    cell = ws_note.cell(1, col)
    cell.font  = copy(HEADER_FONT)
    cell.fill  = copy(HEADER_FILL)
    cell.alignment = copy(CENTER)
ws_note.column_dimensions['A'].width = 22
ws_note.column_dimensions['B'].width = 58

wb_elder.save(path_elder)
print('晚年.xlsx 修正完成（夢境卡方向→選項、填寫說明更新）')

print('\nDone — 全部五個檔案同步完成')
