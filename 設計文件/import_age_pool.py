# -*- coding: utf-8 -*-
"""
通用匯入腳本：讀取五個年齡池 Excel → 輸出 JSON 至 src/
執行方式：python import_age_pool.py

輸出：
  src/phases/02_Lifetime/data/cards.json          （一般牌卡，含全部年齡池）
  src/phases/02_Lifetime/data/dream_cards.json     （夢境卡，含全部年齡池）
  src/phases/02_Lifetime/data/story_chain_cards.json（故事鏈，含全部年齡池）
"""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX_DIR = r'D:\APP\Nice to meet you\設計文件\劇本-正式'
SRC_DIR  = r'D:\APP\Nice to meet you\.claude\worktrees\gracious-lamport\src\phases\02_Lifetime\data'

POOLS = [
    ('童年', 'childhood'),
    ('少年', 'juvenile'),
    ('青年', 'youth'),
    ('中年', 'midlife'),
    ('晚年', 'elder'),
]

EFFECT_MAP = {
    '聲望':     ('surface',    'Reputation'),
    '金錢':     ('surface',    'Money'),
    '裏生命值': ('core',       'Vitality'),
    '穩定':     ('core',       'Stability'),
    '積極':     ('core',       'Drive'),
    '理性':     ('core',       'Logic'),
    '感性':     ('core',       'Emotion'),
    '偽善度':   ('soulToxins', 'Hypocrisy'),
    '冷血度':   ('soulToxins', 'Ruthlessness'),
    '壓抑熵':   ('soulToxins', 'Entropy'),
    '靈魂完整度': ('soulIntegrity', None),
}

# 一般牌卡方向：A/B/C/D → a/b/c/d（同時支援舊格式 上下左右 作為向後相容）
CARD_DIR_MAP = {
    'A': 'a', 'B': 'b', 'C': 'c', 'D': 'd',
    '上': 'a', '下': 'b', '左': 'c', '右': 'd',
}
# 夢境卡與故事鏈：維持 上下左右 → up/down/left/right
FLOW_DIR_MAP = {'上': 'up', '下': 'down', '左': 'left', '右': 'right'}

def parse_effects(row, headers):
    effects = {}
    for zh, (group, en_key) in EFFECT_MAP.items():
        if zh not in headers:
            continue
        val = row.get(zh)
        if val is None or val == '':
            continue
        try:
            val = int(val)
        except (ValueError, TypeError):
            continue
        if val == 0:
            continue
        if group == 'soulIntegrity':
            effects['soulIntegrity'] = val
        else:
            effects.setdefault(group, {})[en_key] = val
    return effects

def parse_condition_str(cond_str):
    """把條件字串解析成 OptionCondition 列表，格式：Key op Value，多條件用 & 連接"""
    if not cond_str:
        return []
    conditions = []
    for part in str(cond_str).split('&'):
        part = part.strip()
        m = re.match(r'(\w+)\s*(>=|<=|>|<|=)\s*(-?\d+(?:\.\d+)?)', part)
        if m:
            conditions.append({'key': m.group(1), 'op': m.group(2), 'value': float(m.group(3))})
    return conditions

def rows_from_sheet(ws):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return headers, rows

# ══════════════════════════════════════════════════════════════
# 一、一般牌卡
# ══════════════════════════════════════════════════════════════
all_cards = {}

for pool_name, pool_en in POOLS:
    path = f'{XLSX_DIR}\\{pool_name}.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb['一般牌卡']
    headers, rows = rows_from_sheet(ws)

    current_id = None
    current_card = None

    for row in rows:
        card_id = str(row.get('卡ID') or '').strip()
        dir_zh  = str(row.get('方向') or '').strip()
        dir_en  = CARD_DIR_MAP.get(dir_zh)
        if not dir_en:
            continue

        # 新卡開始
        if card_id and card_id != current_id:
            current_id = card_id
            situation  = str(row.get('情境') or '').strip()
            tags_raw   = row.get('標籤')
            appear_cond = str(row.get('出現條件') or '').strip() or None
            prob_raw    = row.get('觸發機率')
            prob = int(prob_raw) if prob_raw not in (None, '') else 100

            current_card = {
                'id':        current_id,
                'agePool':   pool_en,
                'situation': situation,
                'tags':      [t.strip() for t in str(tags_raw).split(',') if t and str(tags_raw).strip()] if tags_raw else [],
                'appearCondition': parse_condition_str(appear_cond),
                'probability': prob,
            }
            all_cards[current_id] = current_card

        if current_card is None:
            continue

        # 效果
        effects = parse_effects(row, headers)

        # 後續事件（故事鏈觸發）
        next_event_raw = str(row.get('後續事件') or '').strip()
        next_event = None
        if next_event_raw:
            next_event = {'type': 'storyChain', 'targetId': next_event_raw}

        # 夢境連結
        dream_val_raw = row.get('夢數值')
        dream_val = int(dream_val_raw) if dream_val_raw not in (None, '') else 0
        dream_link = str(row.get('連結夢境卡ID') or '').strip() or None

        # 選項條件
        cond_str = str(row.get('條件') or '').strip()
        conditions = parse_condition_str(cond_str)

        option = {
            'text':    str(row.get('選項文字') or '').strip(),
            'effects': effects,
        }
        if next_event:
            option['nextEvent'] = next_event
        if dream_val:
            option['dreamValue'] = dream_val
        if dream_link:
            option['dreamCardLink'] = dream_link

        # 條件選項
        if conditions:
            base = current_card.get(dir_en)
            if base is None:
                current_card[dir_en] = {'text': '', 'effects': {}}
            current_card[dir_en].setdefault('conditionals', []).append({
                'condition': conditions,
                'text':      option['text'],
                'effects':   effects,
            })
        else:
            current_card[dir_en] = option

print(f'一般牌卡：共 {len(all_cards)} 張')

with open(f'{SRC_DIR}\\cards.json', 'w', encoding='utf-8') as f:
    json.dump({'cards': all_cards}, f, ensure_ascii=False, indent=2)
print('→ cards.json 輸出完成')

# ══════════════════════════════════════════════════════════════
# 二、夢境卡
# ══════════════════════════════════════════════════════════════
all_dreams = {}

for pool_name, pool_en in POOLS:
    path = f'{XLSX_DIR}\\{pool_name}.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb['夢境卡']
    headers, rows = rows_from_sheet(ws)

    current_id = None
    current_dream = None

    for row in rows:
        card_id = str(row.get('卡ID') or '').strip()
        dir_zh  = str(row.get('方向') or '').strip()
        dir_en  = FLOW_DIR_MAP.get(dir_zh)
        if not dir_en:
            continue

        if card_id and card_id != current_id:
            current_id = card_id
            current_dream = {
                'id':        current_id,
                'agePool':   pool_en,
                'situation': str(row.get('夢境情境') or '').strip(),
            }
            all_dreams[current_id] = current_dream

        if current_dream is None:
            continue

        effects = parse_effects(row, headers)
        current_dream[dir_en] = {
            'text':    str(row.get('選項文字') or '').strip(),
            'effects': effects,
        }

print(f'夢境卡：共 {len(all_dreams)} 張')
with open(f'{SRC_DIR}\\dream_cards.json', 'w', encoding='utf-8') as f:
    json.dump({'cards': all_dreams}, f, ensure_ascii=False, indent=2)
print('→ dream_cards.json 輸出完成')

# ══════════════════════════════════════════════════════════════
# 三、故事鏈
# ══════════════════════════════════════════════════════════════
all_chains = {}

for pool_name, pool_en in POOLS:
    path = f'{XLSX_DIR}\\{pool_name}.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb['故事鏈']
    headers, rows = rows_from_sheet(ws)

    current_id = None
    current_chain = None

    for row in rows:
        card_id = str(row.get('卡ID') or '').strip()
        dir_zh  = str(row.get('方向') or '').strip()
        dir_en  = FLOW_DIR_MAP.get(dir_zh)
        if not dir_en:
            continue

        if card_id and card_id != current_id:
            current_id = card_id
            chain_id   = str(row.get('鏈ID') or '').strip()
            current_chain = {
                'id':        current_id,
                'chainId':   chain_id,
                'agePool':   pool_en,
                'situation': str(row.get('情境') or '').strip(),
            }
            all_chains[current_id] = current_chain

        if current_chain is None:
            continue

        ends_day_raw = row.get('結束今日')
        ends_day = str(ends_day_raw).strip().upper() == 'TRUE' if ends_day_raw is not None else False
        next_card = str(row.get('下一張') or '').strip() or None
        effects = parse_effects(row, headers)

        current_chain[dir_en] = {
            'text':    str(row.get('選項文字') or '').strip(),
            'effects': effects,
            'next':    next_card,
            'endsDay': ends_day,
        }

print(f'故事鏈牌：共 {len(all_chains)} 張')
with open(f'{SRC_DIR}\\story_chain_cards.json', 'w', encoding='utf-8') as f:
    json.dump({'cards': all_chains}, f, ensure_ascii=False, indent=2)
print('→ story_chain_cards.json 輸出完成')

print('\nDone — 全部匯入完成')
