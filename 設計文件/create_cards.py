from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active
ws1.title = "卡牌資料"

headers = [
    "卡ID", "年齡池", "情境", "方向", "選項文字",
    "聲望", "金錢", "裏生命值", "穩定", "積極", "理性", "感性",
    "偽善度", "冷血度", "壓抑熵", "靈魂完整度", "靈魂標記", "後續事件"
]

col_widths = {
    "卡ID": 12, "年齡池": 10, "情境": 40, "方向": 6, "選項文字": 30,
    "聲望": 10, "金錢": 10, "裏生命值": 10, "穩定": 10, "積極": 10,
    "理性": 10, "感性": 10, "偽善度": 10, "冷血度": 10, "壓抑熵": 10,
    "靈魂完整度": 10, "靈魂標記": 20, "後續事件": 15
}

header_fill = PatternFill("solid", start_color="FFD700", end_color="FFD700")
white_fill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
gray_fill = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
header_font = Font(name="Arial", bold=True, color="000000")
data_font = Font(name="Arial")
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top")

ws1.append(headers)
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.column_dimensions[get_column_letter(col_idx)].width = col_widths[h]

ws1.row_dimensions[1].height = 20
ws1.freeze_panes = "A2"

cards = [
    {
        "id": "card_001", "pool": "童年",
        "context": "你十二歲。班上總是被欺負的同學，今天又被圍在角落。輪到你要選邊站了。",
        "options": [
            {"dir": "上", "text": "走過去，站到他旁邊", "聲望": -2, "穩定": 3, "積極": 2},
            {"dir": "下", "text": "假裝沒看見，繼續玩你的", "感性": -3, "偽善度": 6},
            {"dir": "左", "text": "試著讓雙方和解，說幾句好話", "聲望": 1, "理性": 2, "感性": 2},
            {"dir": "右", "text": "跟著其他人一起笑。那樣比較安全", "感性": -5, "偽善度": 4, "冷血度": 5},
        ]
    },
    {
        "id": "card_002", "pool": "青年",
        "context": "你二十六歲。一份薪水翻倍的工作機會擺在面前，但必須在四十八小時內決定，沒有時間和任何人商量。",
        "options": [
            {"dir": "上", "text": "接受。人生不等人", "金錢": 15, "積極": 4, "壓抑熵": 3},
            {"dir": "下", "text": "拒絕。現在的地方還有未竟的事", "聲望": 2, "金錢": -5, "穩定": 3},
            {"dir": "左", "text": "詢問能否延後決定期限，爭取時間", "理性": 4, "積極": 2},
            {"dir": "右", "text": "先口頭答應，回去再找理由反悔", "金錢": 10, "穩定": -3, "偽善度": 8},
        ]
    },
    {
        "id": "card_003", "pool": "青年",
        "context": "母親打電話說她不舒服，語氣卻輕描淡寫。你知道她從不說自己痛。",
        "options": [
            {"dir": "上", "text": "立刻請假，當天趕回去", "聲望": -2, "金錢": -5, "感性": 5},
            {"dir": "下", "text": "等週末再回去看她", "感性": -3, "偽善度": 4, "壓抑熵": 5},
            {"dir": "左", "text": "打給家裡其他人，請他們先去看看", "理性": 3, "感性": 2},
            {"dir": "右", "text": "傳訊問她嚴不嚴重。說你最近很忙", "偽善度": 6, "冷血度": 2, "壓抑熵": 4},
        ]
    },
    {
        "id": "card_004", "pool": "中年",
        "context": "你發現，這十年你努力成為的那個人——你其實並不喜歡他。",
        "options": [
            {"dir": "上", "text": "開始做一些改變。小的，但真實的", "積極": 5, "穩定": 3, "壓抑熵": -3},
            {"dir": "下", "text": "告訴自己這就是成熟，是代價。繼續", "穩定": -4, "偽善度": 7, "壓抑熵": 5},
            {"dir": "左", "text": "找一個你信任的人，說出來", "穩定": 2, "感性": 5, "壓抑熵": -4},
            {"dir": "右", "text": "訂了一個旅行計劃。先逃一下再說", "金錢": -10, "積極": -2, "壓抑熵": 3},
        ]
    },
    {
        "id": "card_005", "pool": "中年",
        "context": "有人對你說：「你是我見過最自私的人。」他說的是真的。你知道。",
        "options": [
            {"dir": "上", "text": "沉默，然後離開", "穩定": -2, "冷血度": 2, "壓抑熵": 4},
            {"dir": "下", "text": "反駁，列出你所有的付出", "感性": -4, "偽善度": 8, "冷血度": 3},
            {"dir": "左", "text": "問他能不能說得更具體一點。你想聽", "穩定": 4, "積極": 3, "感性": 5, "壓抑熵": -2},
            {"dir": "右", "text": "笑著說謝謝，心裡把他列為不重要的人", "感性": -6, "偽善度": 5, "冷血度": 6},
        ]
    },
]

stat_cols = ["聲望", "金錢", "裏生命值", "穩定", "積極", "理性", "感性", "偽善度", "冷血度", "壓抑熵", "靈魂完整度"]
header_map = {h: i+1 for i, h in enumerate(headers)}

current_row = 2
for card_idx, card in enumerate(cards):
    fill = white_fill if card_idx % 2 == 0 else gray_fill
    for opt_idx, opt in enumerate(card["options"]):
        row_data = [""] * len(headers)
        if opt_idx == 0:
            row_data[header_map["卡ID"] - 1] = card["id"]
            row_data[header_map["年齡池"] - 1] = card["pool"]
            row_data[header_map["情境"] - 1] = card["context"]
        row_data[header_map["方向"] - 1] = opt["dir"]
        row_data[header_map["選項文字"] - 1] = opt["text"]
        for stat in stat_cols:
            if stat in opt:
                row_data[header_map[stat] - 1] = opt[stat]
        ws1.append(row_data)
        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row=current_row, column=col_idx)
            cell.fill = fill
            cell.font = data_font
            col_name = headers[col_idx - 1]
            if col_name in ("情境", "選項文字"):
                cell.alignment = wrap_align
            else:
                cell.alignment = Alignment(vertical="top")
        row_height = 45 if opt_idx == 0 else 30
        ws1.row_dimensions[current_row].height = row_height
        current_row += 1

ws2 = wb.create_sheet("填寫說明")
ws2.column_dimensions["A"].width = 15
ws2.column_dimensions["B"].width = 60

title_cell = ws2.cell(row=1, column=1, value="卡牌填寫說明")
title_cell.font = Font(name="Arial", bold=True, size=14)
ws2.merge_cells("A1:B1")
ws2.row_dimensions[1].height = 25

field_rows = [
    ("欄位", "說明"),
    ("卡ID", "唯一識別碼，格式建議：card_XXX（三位數字），每張新卡只填第一行"),
    ("年齡池", "童年 / 青年 / 中年 / 老年（每張新卡只填第一行）"),
    ("情境", "顯示在卡牌中央的情境描述（每張新卡只填第一行）"),
    ("方向", "上 / 下 / 左 / 右，每張卡需填滿四行"),
    ("選項文字", "玩家選擇時看到的文字描述"),
    ("聲望", "社會表層數值，可正可負，留空代表此選項不影響此數值"),
    ("金錢", "社會表層數值，可正可負，留空代表此選項不影響此數值"),
    ("裏生命值", "核心數值，可正可負，留空代表不影響"),
    ("穩定", "核心數值（Stability）"),
    ("積極", "核心數值（Drive）"),
    ("理性", "核心數值（Logic）"),
    ("感性", "核心數值（Emotion）"),
    ("偽善度", "靈魂毒素（Hypocrisy），通常為正值（毒素累積）"),
    ("冷血度", "靈魂毒素（Ruthlessness）"),
    ("壓抑熵", "靈魂毒素（Entropy），負值代表毒素淨化"),
    ("靈魂完整度", "直接影響靈魂完整度的特殊效果，通常為負值"),
    ("靈魂標記", "多個標記用逗號分隔，例如：放下,原諒。留空代表無標記"),
    ("後續事件", "故事鏈ID，留空代表正常繼續。例如：story_001"),
]

sub_header_fill = PatternFill("solid", start_color="FFD700", end_color="FFD700")
for r_idx, (field, desc) in enumerate(field_rows, start=2):
    a = ws2.cell(row=r_idx, column=1, value=field)
    b = ws2.cell(row=r_idx, column=2, value=desc)
    b.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[r_idx].height = 30
    if r_idx == 2:
        a.font = Font(name="Arial", bold=True)
        b.font = Font(name="Arial", bold=True)
        a.fill = sub_header_fill
        b.fill = sub_header_fill
    else:
        a.font = Font(name="Arial")
        b.font = Font(name="Arial")

output_path = r"D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
