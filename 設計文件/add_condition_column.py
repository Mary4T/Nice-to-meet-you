# -*- coding: utf-8 -*-
"""
在卡牌資料.xlsx 的「卡牌資料」sheet 插入「條件」欄（情境之後、方向之前）。
現有所有列的條件欄保持空白（向下相容）。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

xlsx_path = r'D:\APP\Nice to meet you\設計文件\卡牌資料.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.worksheets[0]

# 找「方向」欄的位置（1-based）
headers = [cell.value for cell in ws[1]]
# 由於編碼問題，用位置找：方向是第4欄（index 3）
# 先確認
print('目前欄位：', headers)

# 在第4欄（方向）前插入一欄
INSERT_COL = 4  # 在第4欄之前插入，新欄變第4欄

ws.insert_cols(INSERT_COL)

# 設定新欄 header
header_fill = PatternFill('solid', start_color='FFD700', end_color='FFD700')
header_font = Font(name='Arial', bold=True, color='000000')
header_cell = ws.cell(row=1, column=INSERT_COL, value='條件')
header_cell.font = header_font
header_cell.fill = header_fill
header_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions[
    openpyxl.utils.get_column_letter(INSERT_COL)
].width = 22

# 其餘列的新欄設為空白，套用對應行的底色
from openpyxl.styles.fills import PatternFill as PF
white_fill = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
gray_fill  = PatternFill('solid', start_color='F5F5F5', end_color='F5F5F5')
data_font  = Font(name='Arial')

for row_idx in range(2, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=INSERT_COL)
    cell.value = None
    cell.font = data_font
    cell.alignment = Alignment(vertical='top')
    # 繼承該列已有的底色（從相鄰欄取）
    neighbor = ws.cell(row=row_idx, column=INSERT_COL - 1)
    if neighbor.fill and neighbor.fill.fgColor and neighbor.fill.fgColor.rgb != '00000000':
        cell.fill = PatternFill('solid',
                                start_color=neighbor.fill.fgColor.rgb,
                                end_color=neighbor.fill.fgColor.rgb)

wb.save(xlsx_path)
print('Done — 條件欄已插入第 4 欄')
